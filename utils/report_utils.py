"""
Reporting Utility: Exports Test Execution Data to Excel (.xlsx) and PDF (.pdf)
"""

import os
from datetime import datetime
from pathlib import Path
from utils.config import REPORTS_DIR
from utils.db_utils import get_all_results


def export_to_excel(records=None, output_path=None):
    """
    Exports test records to a professional Excel report with summary and details sheets.
    """
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if records is None:
        records = get_all_results()

    if not output_path:
        output_path = REPORTS_DIR / "test_execution_report.xlsx"
    else:
        output_path = Path(output_path)

    if not records:
        print("[WARN] No records available to export to Excel.")
        return str(output_path)

    df = pd.DataFrame(records)

    # Metrics
    total = len(df)
    passed = len(df[df["status"] == "PASSED"])
    failed = len(df[df["status"] == "FAILED"])
    skipped = len(df[df["status"] == "SKIPPED"])
    pass_rate = round((passed / total * 100), 2) if total > 0 else 0.0
    total_time = round(df["duration_seconds"].sum(), 2)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary DataFrame
        summary_data = {
            "Metric": [
                "Total Test Cases Executed",
                "Passed Tests",
                "Failed Tests",
                "Skipped Tests",
                "Overall Pass Rate (%)",
                "Total Execution Duration (s)",
                "Report Generated At",
                "BrowserStack Project",
                "Target Website",
            ],
            "Value": [
                total,
                passed,
                failed,
                skipped,
                f"{pass_rate}%",
                f"{total_time}s",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "webghzyt",
                "https://bugbash.online/",
            ],
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Execution Summary", index=False)

        # Detailed DataFrame
        cols = [
            "id",
            "test_name",
            "category",
            "status",
            "duration_seconds",
            "error_message",
            "executed_at",
            "browser_info",
        ]
        available_cols = [c for c in cols if c in df.columns]
        detail_df = df[available_cols]
        detail_df.to_excel(writer, sheet_name="Test Details", index=False)

    # Open workbook for styling
    wb = openpyxl.load_workbook(output_path)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, bold=True, color="375623")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fail_font = Font(name="Calibri", size=10, bold=True, color="C65911")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.border = thin_border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if str(cell.value) == "PASSED":
                        cell.fill = pass_fill
                        cell.font = pass_font
                    elif str(cell.value) == "FAILED":
                        cell.fill = fail_fill
                        cell.font = fail_font
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    wb.save(output_path)
    print(f"[REPORT] Excel report exported to: {output_path}")
    return str(output_path)


def export_to_pdf(records=None, output_path=None):
    """
    Exports test records to an executive PDF report using ReportLab.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    if records is None:
        records = get_all_results()

    if not output_path:
        output_path = REPORTS_DIR / "test_execution_report.pdf"
    else:
        output_path = Path(output_path)

    if not records:
        print("[WARN] No records available to export to PDF.")
        return str(output_path)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "SubTitleStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#595959"),
        spaceAfter=15,
    )
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )
    cell_bold = ParagraphStyle(
        "CellBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
    )

    story = []

    # Title & Metadata Header
    story.append(Paragraph("BrowserStack Testathon - Test Automation Execution Report", title_style))
    meta_text = (
        f"<b>Project:</b> webghzyt &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Target:</b> https://bugbash.online/ &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Test Management Project:</b> 4102200"
    )
    story.append(Paragraph(meta_text, subtitle_style))

    # Metric calculations
    total = len(records)
    passed = sum(1 for r in records if r.get("status") == "PASSED")
    failed = sum(1 for r in records if r.get("status") == "FAILED")
    skipped = sum(1 for r in records if r.get("status") == "SKIPPED")
    pass_rate = round((passed / total * 100), 2) if total > 0 else 0.0
    total_time = round(sum(r.get("duration_seconds", 0.0) for r in records), 2)

    # Summary Metrics Table
    summary_data = [
        ["Total Tests", "Passed", "Failed", "Skipped", "Pass Rate", "Total Duration"],
        [
            str(total),
            str(passed),
            str(failed),
            str(skipped),
            f"{pass_rate}%",
            f"{total_time}s",
        ],
    ]
    summary_table = Table(summary_data, colWidths=[120, 100, 100, 100, 120, 140])
    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F2F4F7")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D5DD")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])
    )
    story.append(summary_table)
    story.append(Spacer(1, 15))

    # Detailed Test Results Table
    table_headers = ["ID", "Test Case Name", "Category", "Status", "Duration", "Error / Details"]
    table_data = [[Paragraph(f"<b>{h}</b>", cell_bold) for h in table_headers]]

    for r in records:
        status_color = "#2E7D32" if r.get("status") == "PASSED" else "#C62828"
        status_cell = Paragraph(
            f"<font color='{status_color}'><b>{r.get('status')}</b></font>", cell_style
        )
        row = [
            Paragraph(str(r.get("id")), cell_style),
            Paragraph(str(r.get("test_name")), cell_style),
            Paragraph(str(r.get("category", "")), cell_style),
            status_cell,
            Paragraph(f"{r.get('duration_seconds', 0.0):.2f}s", cell_style),
            Paragraph(str(r.get("error_message", ""))[:80], cell_style),
        ]
        table_data.append(row)

    detail_table = Table(
        table_data,
        colWidths=[30, 220, 130, 70, 60, 230],
        repeatRows=1,
    )
    detail_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4053")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D5D8DC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    story.append(detail_table)

    doc.build(story)
    print(f"[REPORT] PDF report exported to: {output_path}")
    return str(output_path)


def generate_all_reports():
    """Generates both Excel and PDF reports from stored database results."""
    records = get_all_results()
    xlsx_path = export_to_excel(records)
    pdf_path = export_to_pdf(records)
    return {"excel": xlsx_path, "pdf": pdf_path}


if __name__ == "__main__":
    generate_all_reports()
